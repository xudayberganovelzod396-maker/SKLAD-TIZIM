"""
SKLAD TIZIM - Ombor Boshqaruv Tizimi
=====================================
Flask-based warehouse management system
Author: Administrator
Version: 2.0
"""

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from functools import wraps
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ==================== APP CONFIGURATION ====================
app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'sklad-tizim-secret-key-2024')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///sklad.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)
app.config['SESSION_COOKIE_HTTPONLY'] = True

db = SQLAlchemy(app)


# ==================== DATABASE MODELS ====================
class User(db.Model):
    """Foydalanuvchilar jadvali"""
    __tablename__ = 'users'
    
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.now)


class Batch(db.Model):
    """Partiyalar jadvali"""
    __tablename__ = 'batches'
    
    id = db.Column(db.Integer, primary_key=True)
    product_name = db.Column(db.String(200), nullable=False)
    batch_code = db.Column(db.String(100), nullable=False)
    quantity = db.Column(db.Integer, nullable=True)
    quantity_sht = db.Column(db.Integer, nullable=True)
    quantity_kg = db.Column(db.Float, nullable=True)
    comment = db.Column(db.String(255), nullable=True)
    location = db.Column(db.String(100), nullable=False)
    status = db.Column(db.String(50), default='ACTIVE')
    is_archived = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    removed_at = db.Column(db.DateTime)
    removed_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    removed_quantity_sht = db.Column(db.Integer, default=0)
    removed_quantity_kg = db.Column(db.Float, default=0.0)
    
    user = db.relationship('User', backref='batches_removed')


class BatchMovement(db.Model):
    """Kirim/Chiqim harakatlari"""
    __tablename__ = 'batch_movements'
    
    id = db.Column(db.Integer, primary_key=True)
    batch_id = db.Column(db.Integer, db.ForeignKey('batches.id'), nullable=False)
    movement_type = db.Column(db.String(10), nullable=False)  # IN / OUT
    quantity_sht = db.Column(db.Integer, default=0)
    quantity_kg = db.Column(db.Float, default=0.0)
    created_at = db.Column(db.DateTime, default=datetime.now)
    
    batch = db.relationship('Batch', backref='movements')


class StockRequest(db.Model):
    """Skladga so'rovlar"""
    __tablename__ = 'stock_requests'

    id = db.Column(db.Integer, primary_key=True)
    product_name = db.Column(db.String(200), nullable=False)
    batch_code = db.Column(db.String(100), nullable=True)
    quantity_sht = db.Column(db.Integer, default=0)
    quantity_kg = db.Column(db.Float, default=0.0)
    comment = db.Column(db.String(255), nullable=True)
    status = db.Column(db.String(20), default='NEW')
    created_at = db.Column(db.DateTime, default=datetime.now)
    seen_at = db.Column(db.DateTime)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))

    user = db.relationship('User', backref='stock_requests')


# ==================== DECORATORS ====================
def login_required(f):
    """Login talab qiluvchi dekorator"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.is_json:
                return jsonify({'error': 'Avtorizatsiya talab qilinadi'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def add_movement(batch, movement_type, qty_sht=0, qty_kg=0.0, created_at=None):
    """Kirim/chiqim harakatini saqlash"""
    movement = BatchMovement(
        batch_id=batch.id,
        movement_type=movement_type,
        quantity_sht=qty_sht or 0,
        quantity_kg=qty_kg or 0.0,
        created_at=created_at or datetime.now()
    )
    db.session.add(movement)


# ==================== MIDDLEWARE ====================
@app.after_request
def set_cache_headers(response):
    """API javoblarini keshlamaslik"""
    if request.path.startswith('/api/'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response


# ==================== AUTH ROUTES ====================
@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login sahifasi"""
    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form
        username = data.get('username')
        password = data.get('password')
        
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            session['user_id'] = user.id
            session.permanent = True
            return jsonify({'success': True, 'message': 'Kirish muvaffaqiyatli!'})
        
        return jsonify({'error': 'Login yoki parol noto\'g\'ri'}), 401
    
    return render_template('login.html')


@app.route('/logout')
def logout():
    """Chiqish"""
    session.clear()
    return redirect(url_for('login'))


# ==================== MAIN ROUTES ====================
@app.route('/')
@login_required
def index():
    """Bosh sahifa"""
    return render_template('index.html')


# ==================== USER API ====================
@app.route('/api/user', methods=['GET'])
@login_required
def get_user():
    """Joriy foydalanuvchi ma'lumotlari"""
    user = db.session.get(User, session['user_id'])
    return jsonify({
        'id': user.id,
        'username': user.username,
        'created_at': user.created_at.strftime('%Y-%m-%d %H:%M')
    })


@app.route('/api/user/password', methods=['POST'])
@login_required
def change_password():
    """Parolni o'zgartirish"""
    data = request.get_json()
    current_password = data.get('current_password')
    new_password = data.get('new_password')
    
    user = db.session.get(User, session['user_id'])
    
    if not check_password_hash(user.password, current_password):
        return jsonify({'error': 'Joriy parol noto\'g\'ri'}), 401
    
    if len(new_password) < 5:
        return jsonify({'error': 'Parol kamida 5 ta belgi bo\'lishi kerak'}), 400
    
    user.password = generate_password_hash(new_password)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Parol muvaffaqiyatli o\'zgartirildi!'})


@app.route('/api/user/activity', methods=['GET'])
@login_required
def get_user_activity():
    """Foydalanuvchi faoliyati statistikasi"""
    user_id = session['user_id']
    total_batches = Batch.query.filter_by(status='ACTIVE').count()
    removed_batches = Batch.query.filter_by(status='REMOVED', removed_by=user_id).count()
    
    return jsonify({
        'total_batches': total_batches,
        'removed_batches': removed_batches,
        'last_active': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })


# ==================== BATCH API ====================
@app.route('/api/batches', methods=['GET'])
@login_required
def get_batches():
    """Barcha partiyalarni olish"""
    batches = Batch.query.order_by(Batch.created_at.desc()).all()
    
    # 0/0 miqdordagi partiyalarni filtrlash
    filtered = [
        b for b in batches
        if not ((b.quantity_sht == 0 or b.quantity_sht is None) and 
                (b.quantity_kg == 0 or b.quantity_kg is None))
    ]
    
    return jsonify([{
        'id': b.id,
        'product_name': b.product_name,
        'batch_code': b.batch_code,
        'quantity': b.quantity,
        'quantity_sht': b.quantity_sht or 0,
        'quantity_kg': b.quantity_kg or 0.0,
        'comment': b.comment,
        'location': b.location,
        'status': b.status,
        'is_archived': b.is_archived,
        'created_at': b.created_at.strftime('%Y-%m-%d %H:%M'),
        'removed_at': b.removed_at.strftime('%Y-%m-%d %H:%M') if b.removed_at else None,
        'removed_by': b.removed_by,
        'removed_quantity_sht': b.removed_quantity_sht or 0,
        'removed_quantity_kg': b.removed_quantity_kg or 0.0
    } for b in filtered])


@app.route('/api/batches', methods=['POST'])
@login_required
def create_batch():
    """Yangi partiya qo'shish"""
    data = request.get_json()
    
    # Majburiy maydonlarni tekshirish
    product_name = data.get('product_name', '').strip()
    batch_code = data.get('batch_code', '').strip()
    location = data.get('location', '').strip()
    
    if not product_name:
        return jsonify({'error': 'Mahsulot nomi kiritilmagan'}), 400
    if not batch_code:
        return jsonify({'error': 'Partiya kodi kiritilmagan'}), 400
    if not location:
        return jsonify({'error': 'Yacheyka kiritilmagan'}), 400
    
    # Miqdorlarni tekshirish
    quantity_sht = data.get('quantity_sht')
    quantity_kg = data.get('quantity_kg')
    
    if quantity_sht is not None:
        try:
            quantity_sht = int(quantity_sht)
            if quantity_sht < 0:
                return jsonify({'error': 'Dona miqdor manfiy bo\'lishi mumkin emas'}), 400
        except (ValueError, TypeError):
            return jsonify({'error': 'Noto\'g\'ri dona miqdor'}), 400
    
    if quantity_kg is not None:
        try:
            quantity_kg = float(quantity_kg)
            if quantity_kg < 0:
                return jsonify({'error': 'Kg miqdor manfiy bo\'lishi mumkin emas'}), 400
        except (ValueError, TypeError):
            return jsonify({'error': 'Noto\'g\'ri kg miqdor'}), 400
    
    # Kamida bitta miqdor kiritilishi kerak
    if (quantity_sht is None or quantity_sht == 0) and (quantity_kg is None or quantity_kg == 0):
        return jsonify({'error': 'Kamida bitta miqdor kiritilishi kerak (dona yoki kg)'}), 400
    
    quantity = quantity_sht or quantity_kg or 0
    
    batch = Batch(
        product_name=product_name,
        batch_code=batch_code,
        quantity=quantity,
        quantity_sht=quantity_sht,
        quantity_kg=quantity_kg,
        comment=data.get('comment', '').strip(),
        location=location,
        status='ACTIVE',
        is_archived=False,
        removed_quantity_sht=0,
        removed_quantity_kg=0.0
    )
    
    db.session.add(batch)
    db.session.commit()

    add_movement(
        batch,
        movement_type='IN',
        qty_sht=quantity_sht or 0,
        qty_kg=quantity_kg or 0.0,
        created_at=batch.created_at
    )
    db.session.commit()
    
    return jsonify({'success': True, 'batch_id': batch.id}), 201


@app.route('/api/batches/<int:batch_id>/remove', methods=['PUT'])
@login_required
def remove_batch(batch_id):
    """Partiyani chiqarish"""
    batch = db.session.get(Batch, batch_id)
    
    if not batch:
        return jsonify({'error': 'Partiya topilmadi'}), 404
    
    if batch.is_archived:
        return jsonify({'error': 'Arxivlangan partiya o\'zgartirilishi mumkin emas'}), 403
    
    data = request.get_json(silent=True) or {}
    qty_sht = data.get('quantity_sht')
    qty_kg = data.get('quantity_kg')
    
    if qty_sht is None and qty_kg is None:
        return jsonify({'error': 'Miqdor kiritilmadi'}), 400
    
    # Dona miqdorini tekshirish
    if qty_sht is not None:
        try:
            qty_sht = int(qty_sht)
        except (ValueError, TypeError):
            return jsonify({'error': 'Noto\'g\'ri dona miqdor'}), 400
        
        if qty_sht < 0 or (batch.quantity_sht is not None and qty_sht > batch.quantity_sht):
            return jsonify({'error': f'Dona miqdor 0 dan {batch.quantity_sht} gacha bo\'lishi kerak'}), 400
    
    # Kg miqdorini tekshirish
    if qty_kg is not None:
        try:
            qty_kg = float(qty_kg)
        except (ValueError, TypeError):
            return jsonify({'error': 'Noto\'g\'ri kg miqdor'}), 400
        
        if qty_kg < 0 or (batch.quantity_kg is not None and qty_kg > batch.quantity_kg):
            return jsonify({'error': f'Kg miqdor 0 dan {batch.quantity_kg} gacha bo\'lishi kerak'}), 400
    
    # Qisman chiqarish - qolgan miqdorni hisoblash
    remaining_sht = batch.quantity_sht
    remaining_kg = batch.quantity_kg
    
    if qty_sht is not None and remaining_sht is not None:
        remaining_sht = remaining_sht - qty_sht
    if qty_kg is not None and remaining_kg is not None:
        remaining_kg = remaining_kg - qty_kg
    
    # Chiqarilgan miqdorlarni qo'shib saqlash (accumulated)
    if qty_sht is not None:
        batch.removed_quantity_sht = (batch.removed_quantity_sht or 0) + qty_sht
    if qty_kg is not None:
        batch.removed_quantity_kg = (batch.removed_quantity_kg or 0) + qty_kg
    
    # Qolgan miqdorni yangilash
    batch.quantity_sht = remaining_sht
    batch.quantity_kg = remaining_kg
    
    # Agar hammasi chiqarilgan bo'lsa
    if (remaining_sht is None or remaining_sht <= 0) and (remaining_kg is None or remaining_kg <= 0):
        batch.status = 'REMOVED'
        batch.removed_at = datetime.now()
        batch.removed_by = session['user_id']
        batch.quantity_sht = 0
        batch.quantity_kg = 0
        batch.is_archived = True

    add_movement(
        batch,
        movement_type='OUT',
        qty_sht=qty_sht or 0,
        qty_kg=qty_kg or 0.0,
        created_at=datetime.now()
    )
    
    db.session.commit()
    
    return jsonify({'success': True})


# ==================== SEARCH API ====================
@app.route('/api/search', methods=['GET'])
@login_required
def search():
    """Partiyalarni qidirish"""
    query = request.args.get('q', '').lower()
    
    if not query:
        return jsonify([])
    
    results = Batch.query.filter(
        Batch.status == 'ACTIVE',
        db.or_(
            Batch.batch_code.ilike(f'%{query}%'),
            Batch.product_name.ilike(f'%{query}%'),
            Batch.location.ilike(f'%{query}%')
        )
    ).all()
    
    return jsonify([{
        'id': b.id,
        'product_name': b.product_name,
        'batch_code': b.batch_code,
        'quantity': b.quantity,
        'quantity_sht': b.quantity_sht or 0,
        'quantity_kg': b.quantity_kg or 0.0,
        'location': b.location,
        'is_archived': b.is_archived,
        'created_at': b.created_at.strftime('%Y-%m-%d %H:%M')
    } for b in results])


@app.route('/api/batches/search', methods=['GET'])
@login_required
def search_batches():
    """Partiyalarni sahifalash bilan qidirish"""
    query = request.args.get('q', '').strip()
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('page_size', 7))
    
    if not query:
        return jsonify({'results': [], 'total': 0})
    
    q = f"%{query.lower()}%"
    batches_query = Batch.query.filter(
        Batch.status == 'ACTIVE',
        db.or_(
            Batch.product_name.ilike(q),
            Batch.batch_code.ilike(q),
            Batch.location.ilike(q)
        )
    ).order_by(Batch.created_at.desc())
    
    total = batches_query.count()
    batches = batches_query.offset((page - 1) * page_size).limit(page_size).all()
    
    results = [{
        'id': b.id,
        'product_name': b.product_name,
        'batch_code': b.batch_code,
        'quantity': b.quantity,
        'quantity_sht': b.quantity_sht or 0,
        'quantity_kg': b.quantity_kg or 0.0,
        'comment': b.comment,
        'location': b.location,
        'status': b.status,
        'is_archived': b.is_archived,
        'created_at': b.created_at.strftime('%Y-%m-%d %H:%M'),
        'removed_at': b.removed_at.strftime('%Y-%m-%d %H:%M') if b.removed_at else None,
        'removed_by': b.removed_by
    } for b in batches]
    
    return jsonify({'results': results, 'total': total})


@app.route('/api/batches/by-code', methods=['GET'])
@login_required
def get_batch_by_code():
    """Partiya kodi bo'yicha ma'lumot olish"""
    code = request.args.get('code', '').strip()
    if not code:
        return jsonify({'error': 'Partiya kodi kiritilmagan'}), 400

    batches = Batch.query.filter(
        Batch.status == 'ACTIVE',
        Batch.batch_code == code
    ).order_by(Batch.created_at.desc()).all()

    if not batches:
        return jsonify({'error': 'Partiya topilmadi'}), 404

    total_sht = sum(b.quantity_sht or 0 for b in batches)
    total_kg = sum(b.quantity_kg or 0.0 for b in batches)
    product_names = list({b.product_name for b in batches})
    product_name = product_names[0] if len(product_names) == 1 else product_names[0]

    return jsonify({
        'product_name': product_name,
        'quantity_sht': total_sht,
        'quantity_kg': total_kg,
        'items': [{
            'id': b.id,
            'product_name': b.product_name,
            'batch_code': b.batch_code,
            'quantity_sht': b.quantity_sht or 0,
            'quantity_kg': b.quantity_kg or 0.0,
            'location': b.location,
            'created_at': b.created_at.strftime('%Y-%m-%d %H:%M')
        } for b in batches]
    })


# ==================== WAREHOUSE STATUS API ====================
@app.route('/api/rows_matrix_status')
@login_required
def rows_matrix_status():
    """Ombor matritsa holati"""
    sectors = ['A', 'B', 'C']
    rows = 9
    cells = 4
    matrix = {}
    
    for sector in sectors:
        matrix[sector] = []
        for row in range(1, rows + 1):
            row_cells = []
            for cell in range(1, cells + 1):
                loc = f"{sector}-{row}-{cell}"
                batch = Batch.query.filter_by(location=loc).filter(
                    ((Batch.quantity_sht != None) & (Batch.quantity_sht > 0)) |
                    ((Batch.quantity_kg != None) & (Batch.quantity_kg > 0))
                ).first()
                row_cells.append('busy' if batch else 'free')
            matrix[sector].append(row_cells)
    
    return jsonify(matrix)


# ==================== ARCHIVE API ====================
@app.route('/api/archive', methods=['GET'])
@login_required
def get_archive():
    """Arxiv ma'lumotlarini olish"""
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    day = request.args.get('day')
    
    movements_query = BatchMovement.query.join(Batch)
    
    # Sana bo'yicha filtrlash
    if start_date_str or end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None
        except ValueError:
            return jsonify({'error': 'Sana noto\'g\'ri formatda'}), 400
        if start_date and end_date and end_date < start_date:
            return jsonify({'error': 'Sana oralig\'i noto\'g\'ri'}), 400
        if start_date:
            movements_query = movements_query.filter(BatchMovement.created_at >= start_date)
        if end_date:
            end_exclusive = end_date + timedelta(days=1)
            movements_query = movements_query.filter(BatchMovement.created_at < end_exclusive)
    elif day:
        try:
            start_date = datetime.strptime(day, '%Y-%m-%d')
            end_date = start_date + timedelta(days=1)
        except ValueError:
            return jsonify({'error': 'Sana noto\'g\'ri formatda'}), 400
        movements_query = movements_query.filter(
            BatchMovement.created_at >= start_date,
            BatchMovement.created_at < end_date
        )
    elif year and month:
        start_date = datetime(year, month, 1)
        end_date = datetime(year + 1, 1, 1) if month == 12 else datetime(year, month + 1, 1)
        movements_query = movements_query.filter(
            BatchMovement.created_at >= start_date,
            BatchMovement.created_at < end_date
        )
    elif year:
        start_date = datetime(year, 1, 1)
        end_date = datetime(year + 1, 1, 1)
        movements_query = movements_query.filter(
            BatchMovement.created_at >= start_date,
            BatchMovement.created_at < end_date
        )
    elif month:
        current_year = datetime.now().year
        start_date = datetime(current_year, month, 1)
        end_date = datetime(current_year + 1, 1, 1) if month == 12 else datetime(current_year, month + 1, 1)
        movements_query = movements_query.filter(
            BatchMovement.created_at >= start_date,
            BatchMovement.created_at < end_date
        )
    
    incoming_movements = movements_query.filter(BatchMovement.movement_type == 'IN').all()
    outgoing_movements = movements_query.filter(BatchMovement.movement_type == 'OUT').all()
    
    def aggregate_movements(movements):
        aggregated = {}
        for m in movements:
            b = m.batch
            if not b:
                continue
            key = (b.batch_code, b.product_name)
            if key not in aggregated:
                aggregated[key] = {
                    'product_name': b.product_name,
                    'batch_code': b.batch_code,
                    'quantity_sht': 0,
                    'quantity_kg': 0.0
                }
            aggregated[key]['quantity_sht'] += m.quantity_sht or 0
            aggregated[key]['quantity_kg'] += m.quantity_kg or 0.0
        return list(aggregated.values())
    
    return jsonify({
        'incoming': aggregate_movements(incoming_movements),
        'outgoing': aggregate_movements(outgoing_movements)
    })


# ==================== STOCK REQUEST API ====================
@app.route('/api/requests', methods=['GET'])
@login_required
def get_stock_requests():
    """Sklad so'rovlarini olish"""
    status = request.args.get('status')
    query = StockRequest.query.order_by(StockRequest.created_at.desc())
    if status:
        if status == 'COMPLETED':
            query = query.filter(StockRequest.status.in_(['DONE', 'FAILED']))
        else:
            query = query.filter(StockRequest.status == status)
    requests_list = query.all()

    batch_codes = [r.batch_code for r in requests_list if r.batch_code]
    location_map = {}
    if batch_codes:
        batches = (Batch.query
                   .filter(Batch.batch_code.in_(batch_codes))
                   .order_by(Batch.created_at.desc())
                   .all())
        for b in batches:
            if b.batch_code not in location_map:
                location_map[b.batch_code] = b.location

    return jsonify([{
        'id': r.id,
        'product_name': r.product_name,
        'batch_code': r.batch_code,
        'location': location_map.get(r.batch_code) if r.batch_code else None,
        'quantity_sht': r.quantity_sht or 0,
        'quantity_kg': r.quantity_kg or 0.0,
        'comment': r.comment,
        'status': r.status,
        'created_at': (r.seen_at if r.status in ['DONE', 'FAILED'] and r.seen_at else r.created_at).strftime('%Y-%m-%d %H:%M'),
        'seen_at': r.seen_at.strftime('%Y-%m-%d %H:%M') if r.seen_at else None,
        'created_by': r.created_by
    } for r in requests_list])


@app.route('/api/requests', methods=['POST'])
@login_required
def create_stock_request():
    """Skladga so'rov yaratish"""
    data = request.get_json() or {}
    product_name = (data.get('product_name') or '').strip()
    batch_code = (data.get('batch_code') or '').strip() or None
    comment = (data.get('comment') or '').strip() or None

    if not product_name:
        return jsonify({'error': 'Mahsulot nomi kiritilmagan'}), 400

    qty_sht = data.get('quantity_sht')
    qty_kg = data.get('quantity_kg')

    if qty_sht is not None:
        try:
            qty_sht = int(qty_sht)
        except (ValueError, TypeError):
            return jsonify({'error': 'Noto\'g\'ri dona miqdor'}), 400
        if qty_sht < 0:
            return jsonify({'error': 'Dona miqdor manfiy bo\'lishi mumkin emas'}), 400

    if qty_kg is not None:
        try:
            qty_kg = float(qty_kg)
        except (ValueError, TypeError):
            return jsonify({'error': 'Noto\'g\'ri kg miqdor'}), 400
        if qty_kg < 0:
            return jsonify({'error': 'Kg miqdor manfiy bo\'lishi mumkin emas'}), 400

    qty_sht = qty_sht or 0
    qty_kg = qty_kg or 0.0
    if qty_sht == 0 and qty_kg == 0:
        return jsonify({'error': 'Kamida bitta miqdor kiriting'}), 400

    new_request = StockRequest(
        product_name=product_name,
        batch_code=batch_code,
        quantity_sht=qty_sht,
        quantity_kg=qty_kg,
        comment=comment,
        status='NEW',
        created_by=session['user_id']
    )
    db.session.add(new_request)
    db.session.commit()

    return jsonify({'success': True, 'id': new_request.id})


@app.route('/api/requests/<int:req_id>/seen', methods=['PUT'])
@login_required
def mark_request_seen(req_id):
    """So'rovni ko'rildi deb belgilash"""
    req = db.session.get(StockRequest, req_id)
    if not req:
        return jsonify({'error': 'So\'rov topilmadi'}), 404
    req.status = 'SEEN'
    req.seen_at = datetime.now()
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/requests/<int:req_id>/done', methods=['PUT'])
@login_required
def mark_request_done(req_id):
    """So'rovni bajarildi deb belgilash"""
    req = db.session.get(StockRequest, req_id)
    if not req:
        return jsonify({'error': 'So\'rov topilmadi'}), 404
    req.status = 'DONE'
    req.seen_at = datetime.now()
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/requests/<int:req_id>/failed', methods=['PUT'])
@login_required
def mark_request_failed(req_id):
    """So'rovni bajarilmadi deb belgilash"""
    req = db.session.get(StockRequest, req_id)
    if not req:
        return jsonify({'error': 'So\'rov topilmadi'}), 404
    req.status = 'FAILED'
    req.seen_at = datetime.now()
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/archive/export', methods=['GET'])
@login_required
def export_archive_excel():
    """Arxiv ma'lumotlarini Excelga export qilish"""
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    day = request.args.get('day')
    search = (request.args.get('search') or '').strip().lower()

    movements_query = BatchMovement.query.join(Batch)

    if start_date_str or end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None
        except ValueError:
            return jsonify({'error': 'Sana noto\'g\'ri formatda'}), 400
        if start_date and end_date and end_date < start_date:
            return jsonify({'error': 'Sana oralig\'i noto\'g\'ri'}), 400
        if start_date:
            movements_query = movements_query.filter(BatchMovement.created_at >= start_date)
        if end_date:
            end_exclusive = end_date + timedelta(days=1)
            movements_query = movements_query.filter(BatchMovement.created_at < end_exclusive)
    elif day:
        try:
            start_date = datetime.strptime(day, '%Y-%m-%d')
            end_date = start_date + timedelta(days=1)
        except ValueError:
            return jsonify({'error': 'Sana noto\'g\'ri formatda'}), 400
        movements_query = movements_query.filter(
            BatchMovement.created_at >= start_date,
            BatchMovement.created_at < end_date
        )
    elif year and month:
        start_date = datetime(year, month, 1)
        end_date = datetime(year + 1, 1, 1) if month == 12 else datetime(year, month + 1, 1)
        movements_query = movements_query.filter(
            BatchMovement.created_at >= start_date,
            BatchMovement.created_at < end_date
        )
    elif year:
        start_date = datetime(year, 1, 1)
        end_date = datetime(year + 1, 1, 1)
        movements_query = movements_query.filter(
            BatchMovement.created_at >= start_date,
            BatchMovement.created_at < end_date
        )
    elif month:
        current_year = datetime.now().year
        start_date = datetime(current_year, month, 1)
        end_date = datetime(current_year + 1, 1, 1) if month == 12 else datetime(current_year, month + 1, 1)
        movements_query = movements_query.filter(
            BatchMovement.created_at >= start_date,
            BatchMovement.created_at < end_date
        )

    incoming_movements = movements_query.filter(BatchMovement.movement_type == 'IN').all()
    outgoing_movements = movements_query.filter(BatchMovement.movement_type == 'OUT').all()

    def aggregate_movements(movements):
        aggregated = {}
        for m in movements:
            b = m.batch
            if not b:
                continue
            if search:
                code = (b.batch_code or '').lower()
                name = (b.product_name or '').lower()
                if search not in code and search not in name:
                    continue
            key = (b.batch_code, b.product_name)
            if key not in aggregated:
                aggregated[key] = {
                    'product_name': b.product_name,
                    'batch_code': b.batch_code,
                    'quantity_sht': 0,
                    'quantity_kg': 0.0
                }
            aggregated[key]['quantity_sht'] += m.quantity_sht or 0
            aggregated[key]['quantity_kg'] += m.quantity_kg or 0.0
        return list(aggregated.values())

    incoming = aggregate_movements(incoming_movements)
    outgoing = aggregate_movements(outgoing_movements)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Arxiv'

    header_fill_in = PatternFill(start_color='DFF5E1', end_color='DFF5E1', fill_type='solid')
    header_fill_out = PatternFill(start_color='FADDE2', end_color='FADDE2', fill_type='solid')
    title_fill_in = PatternFill(start_color='2DBE60', end_color='2DBE60', fill_type='solid')
    title_fill_out = PatternFill(start_color='E04A6A', end_color='E04A6A', fill_type='solid')
    title_font = Font(color='FFFFFF', bold=True, size=12)
    header_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    thin = Side(border_style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 12

    # Titles
    ws.merge_cells('A1:D1')
    ws['A1'] = '📥 Приход'
    ws['A1'].fill = title_fill_in
    ws['A1'].font = title_font
    ws['A1'].alignment = center

    ws.merge_cells('F1:I1')
    ws['F1'] = '📤 Расход'
    ws['F1'].fill = title_fill_out
    ws['F1'].font = title_font
    ws['F1'].alignment = center

    def write_table(start_col, title_fill, data):
        headers = ['Товар', 'Партия', 'Шт', 'Кг']
        start_row = 2
        for idx, h in enumerate(headers):
            cell = ws.cell(row=start_row, column=start_col + idx, value=h)
            cell.font = header_font
            cell.fill = title_fill
            cell.alignment = center
            cell.border = border

        row = start_row + 1
        total_sht = 0
        total_kg = 0.0
        for item in data:
            ws.cell(row=row, column=start_col, value=item['product_name']).alignment = left
            ws.cell(row=row, column=start_col + 1, value=item['batch_code']).alignment = left
            ws.cell(row=row, column=start_col + 2, value=item['quantity_sht'])
            ws.cell(row=row, column=start_col + 3, value=float(item['quantity_kg']))
            for c in range(start_col, start_col + 4):
                ws.cell(row=row, column=c).border = border
            total_sht += item['quantity_sht']
            total_kg += float(item['quantity_kg'])
            row += 1

        # Totals row
        ws.cell(row=row, column=start_col, value='Итого:').font = header_font
        ws.cell(row=row, column=start_col + 2, value=total_sht).font = header_font
        ws.cell(row=row, column=start_col + 3, value=total_kg).font = header_font
        for c in range(start_col, start_col + 4):
            ws.cell(row=row, column=c).fill = title_fill
            ws.cell(row=row, column=c).border = border

    write_table(1, header_fill_in, incoming)
    write_table(6, header_fill_out, outgoing)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    period = ''
    if day:
        period = day
    elif year and month:
        period = f"{year}-{str(month).zfill(2)}"
    elif year:
        period = f"{year}"
    elif month:
        period = f"{str(month).zfill(2)}"

    filename = f"архив_{period or 'all'}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ==================== REPORT API ====================
@app.route('/api/report', methods=['GET'])
@login_required
def report():
    """Kirim/chiqim hisoboti"""
    start_date = request.args.get('start')
    end_date = request.args.get('end')
    
    try:
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
    except (ValueError, TypeError):
        return jsonify({'error': 'Sanalar noto\'g\'ri formatda'}), 400
    
    movements = BatchMovement.query.filter(
        BatchMovement.created_at >= start,
        BatchMovement.created_at <= end
    ).all()
    
    incoming = [m for m in movements if m.movement_type == 'IN']
    outgoing = [m for m in movements if m.movement_type == 'OUT']
    
    kirim_count = len({m.batch_id for m in incoming})
    kirim_kg = sum(m.quantity_kg or 0 for m in incoming)
    kirim_sht = sum(m.quantity_sht or 0 for m in incoming)
    
    chiqim_count = len({m.batch_id for m in outgoing})
    chiqim_kg = sum(m.quantity_kg or 0 for m in outgoing)
    chiqim_sht = sum(m.quantity_sht or 0 for m in outgoing)
    
    return jsonify({
        'kirim': {'partiya': kirim_count, 'kg': kirim_kg, 'sht': kirim_sht},
        'chiqim': {'partiya': chiqim_count, 'kg': chiqim_kg, 'sht': chiqim_sht}
    })


# ==================== ERROR HANDLERS ====================
@app.errorhandler(404)
def not_found(e):
    """404 xatosi"""
    if request.is_json or request.path.startswith('/api/'):
        return jsonify({'error': 'Topilmadi'}), 404
    return render_template('404.html'), 404


@app.errorhandler(500)
def server_error(e):
    """500 xatosi"""
    return jsonify({'error': 'Server xatosi'}), 500


# ==================== DATABASE INITIALIZATION ====================
def init_db():
    """Ma'lumotlar bazasini yaratish"""
    with app.app_context():
        db.create_all()
        
        # Admin foydalanuvchi yaratish
        if not User.query.filter_by(username='admin').first():
            admin = User(
                username='admin',
                password=generate_password_hash('admin123')
            )
            db.session.add(admin)
        
        # Oddiy foydalanuvchi yaratish
        if not User.query.filter_by(username='user').first():
            user = User(
                username='user',
                password=generate_password_hash('user123')
            )
            db.session.add(user)
        
        db.session.commit()

        # Mavjud partiyalar uchun kirim/chiqim tarixini to'ldirish (best-effort)
        existing_in = {m.batch_id for m in BatchMovement.query.filter_by(movement_type='IN').all()}
        existing_out = {m.batch_id for m in BatchMovement.query.filter_by(movement_type='OUT').all()}
        
        for b in Batch.query.all():
            if b.id not in existing_in:
                total_sht = (b.quantity_sht or 0) + (b.removed_quantity_sht or 0)
                total_kg = (b.quantity_kg or 0.0) + (b.removed_quantity_kg or 0.0)
                if total_sht > 0 or total_kg > 0:
                    add_movement(
                        b,
                        movement_type='IN',
                        qty_sht=total_sht,
                        qty_kg=total_kg,
                        created_at=b.created_at
                    )
            if b.removed_at and b.id not in existing_out:
                if (b.removed_quantity_sht or 0) > 0 or (b.removed_quantity_kg or 0.0) > 0:
                    add_movement(
                        b,
                        movement_type='OUT',
                        qty_sht=b.removed_quantity_sht or 0,
                        qty_kg=b.removed_quantity_kg or 0.0,
                        created_at=b.removed_at
                    )
        
        db.session.commit()


# ==================== RUN APPLICATION ====================
if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='127.0.0.1', port=5000)
